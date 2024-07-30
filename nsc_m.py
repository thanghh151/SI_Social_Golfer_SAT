from pysat.solvers import Glucose3, Solver
from prettytable import PrettyTable
from threading import Timer
import datetime
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

num_weeks: int  # number of weeks
players_per_group: int  # players per group
num_groups: int  # number of groups
num_players: int  # players per group * number of groups
id_variable: int
time_budget = 600
all_clauses = []
id_counter = 0

sat_solver: Solver
enable_pythonsat = True
enable_kissat = False

online_path = ''
log_file = open(online_path + 'console.log', 'a')

# Define a custom print function that writes to both console and log file
def print_to_console_and_log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file = log_file, **kwargs)
    log_file.flush()

def is_prime(x: int) -> bool:
    if x < 2: return False
    for i in range(2, int(x ** 0.5) + 1):
        if x % i == 0: return False
    return True

def generate_all_clauses():
    ensure_golfer_plays_exactly_once_per_week()
    ensure_group_contains_exactly_p_players()
    ensure_no_repeated_players_in_groups()
    symmetry_breaking_1()
    symmetry_breaking_2()
    # symmetry_breaking_7()
    global num_groups, players_per_group, num_weeks
    if num_groups == players_per_group and num_weeks == players_per_group + 1:
        symmetry_breaking_8()
        symmetry_breaking_9()
        if num_groups <= 7 or num_groups == 9:
            symmetry_breaking_10()
            symmetry_breaking_11()
            symmetry_breaking_12()
            if num_groups == 9: bonus_9_9_10()
        else:
            if is_prime(num_groups):
                symmetry_breaking_13_14()
                build_gs1()
            if num_groups == 8:
                # not_enough_10()
                symmetry_breaking_13_14()
                symmetry_breaking_15()
                symmetry_breaking_16()

def plus_clause(clause):
    if enable_pythonsat: sat_solver.add_clause(clause)
    if enable_kissat: all_clauses.append(clause)

# (EO) Using binomial
def exactly_one(var: list[int]):
    n = len(var)
    assert n == num_groups

    # (1): (ALO)
    clause = []
    for i in range(0, n):
        clause.append(var[i])
    plus_clause(clause)

    # (2): (AMO)
    for i in range (0, n):
        for j in range (i + 1, n):
            plus_clause([-var[i], -var[j]])

# Every player plays exactly once a week
# x_w_g (1)
def ensure_golfer_plays_exactly_once_per_week():
    for player in range(min(num_groups, players_per_group) + 1, num_players + 1):
        for week in range(2, num_weeks + 1):
            list = []
            for group in range(1, num_groups + 1):
                list.append(get_variable(player, group, week))
            exactly_one(list)

# (EK) Using New Sequential encounter (NSC)
def exactly_k(var: list[int], k):
    global id_variable
    n = len(var) - 1
    assert n == num_players
    map_register = [[0 for j in range(k + 1)] for i in range(n)]
    for i in range(1, n):
        for j in range(1, min(i, k) + 1):
            id_variable += 1
            map_register[i][j] = id_variable

    # (1): If a bit is true, the first bit of the corresponding register is true
    for i in range(1, n):
        plus_clause([-var[i], map_register[i][1]])

    # (2): R[i - 1][j] = 1, R[i][j] = 1;
    for i in range(2, n):
        for j in range(1, min(i - 1, k) + 1):
            plus_clause([-map_register[i - 1][j], map_register[i][j]])

    # (3): If bit i is on and R[i - 1][j - 1] = 1, R[i][j] = 1;
    for i in range(2, n):
        for j in range(2, min(i, k) + 1):
            plus_clause([-var[i], -map_register[i - 1][j - 1], map_register[i][j]])

    # (4): If bit i is off and R[i - 1][j] = 0, R[i][j] = 0;
    for i in range(2, n):
        for j in range(1, min(i - 1, k) + 1):
            plus_clause([var[i], map_register[i - 1][j], -map_register[i][j]])

    # (5): If bit i is off, R[i][i] = 0;
    for i in range(1, k + 1):
        plus_clause([var[i], -map_register[i][i]])

    # (6): If R[i - 1][j - 1] = 0, R[i][j] = 0;
    for i in range(2, n):
        for j in range(2, min(i, k) + 1):
            plus_clause([map_register[i - 1][j - 1], -map_register[i][j]])

    # (7): (At least k) R[n - 1][k] = 1 or (n-th bit is true and R[n - 1][k - 1] = 1)
    plus_clause([map_register[n - 1][k], var[n]])
    plus_clause([map_register[n - 1][k], map_register[n - 1][k - 1]])
    # plus_clause([map_register[n - 1][k - 1]])

    # (8): (At most k) If i-th bit is true, R[i - 1][k] = 0;
    for i in range(k + 1, n + 1):
        plus_clause([-var[i], -map_register[i - 1][k]])

# A group contains exactly p players
# w_g_x (2)
def ensure_group_contains_exactly_p_players():
    for week in range(2, num_weeks + 1):
        for group in range(1, num_groups + 1):
            list = [-1]
            for player in range(1, num_players + 1):
                list.append(get_variable(player, group, week))
            exactly_k(list, players_per_group)

# Ensures that no players are repeated in the same group across different weeks and groups.
# w_g_x_x_g_w (3)
def ensure_no_repeated_players_in_groups():
    if num_weeks == 1: return
    M = [0 for i in range(num_weeks)] # SC encoding

    def at_most_one(p1, p2):
        global id_variable
        for i in range(0, num_weeks):
            M[i] = id_variable + 1
            id_variable += 1
        
        # (1): M[0] = 0
        plus_clause([-M[0]])

        # (2): If M[i - 1] = 1, M[i] = 1
        for i in range(1, num_weeks): plus_clause([-M[i - 1], M[i]])

        # (3): If p1 and p2 meet in week i, M[i] = 1
        for i in range(1, num_weeks):
            for g in range(1, num_groups + 1):
                plus_clause([-get_variable(p1, g, i), -get_variable(p2, g, i), M[i]])

        # (4): If M[i - 1] = 0 and (p1 and p2) don't meet in week i, M[i] = 0
        for i in range(1, num_weeks):
            for g in range(1, num_groups + 1):
                plus_clause([M[i - 1], -get_variable(p1, g, i), get_variable(p2, g, i), -M[i]])

        # (5): If M[i - 1] = 1, (p1 and p2) don't meet in week i
        for i in range(2, num_weeks + 1):
            for g in range(1, num_groups + 1):
                plus_clause([-M[i - 1], -get_variable(p1, g, i), -get_variable(p2, g, i)])

    for p1 in range(1, num_players + 1):
        for p2 in range(p1 + 1, num_players + 1): at_most_one(p1, p2)


# SB1: The first week order is [1, 2, 3, ... x]
def symmetry_breaking_1():
    for player in range(1, num_players + 1):
        right_group = (player - 1) // players_per_group + 1
        for group in range(1, num_groups + 1):
            if group == right_group:
                plus_clause([get_variable(player, group, 1)])
            else:
                plus_clause([-get_variable(player, group, 1)])

# SB2: From week 2, first p players belong to p groups
def symmetry_breaking_2():
    for week in range(2, num_weeks + 1):
        for player in range(1, min(num_groups, players_per_group) + 1):
            for group in range(1, num_groups + 1):
                if group == player:
                    plus_clause([get_variable(player, group, week)])
                else:
                    plus_clause([-get_variable(player, group, week)])

# SB7
def symmetry_breaking_7():
    for week in range(2, num_weeks + 1):
        for group in range(1, num_groups + 1):
            for golfer1 in range(1, num_players + 1):
                for golfer2 in range(golfer1 + 1, num_players + 1):
                        if ((golfer1 - 1) // players_per_group == (golfer2 - 1) // players_per_group):
                            clause = [-get_variable(golfer1, group, week),
                                      -get_variable(golfer2, group, week)]
                            plus_clause(clause)

# SB8
def symmetry_breaking_8():
    for group in range(1, num_groups + 1):
        for week in range(2, num_weeks + 1):
            for other_week in range(week + 1, num_weeks + 1):
                for player in range(players_per_group + 1, num_players + 1):
                    plus_clause([-get_variable(player, group, week), -get_variable(player, group, other_week)])

# SB9
def symmetry_breaking_9():
    for player in range(players_per_group + 1, num_players + 1):
        plus_clause([get_variable(player, (player - 1) % num_groups + 1, 2)])

# SB10
def symmetry_breaking_10():
    diff = players_per_group - 1
    for group in range(1, num_groups + 1):
        for week in range(2, num_weeks + 1):
            for player in range(players_per_group + 1, 2 * players_per_group + 1):
                if (player - diff != week):
                    plus_clause([get_variable(player, group, week), -get_variable(week + diff, group, player - diff)])

# The [p + 1] column is [2, 1, 2, 3, ... , g]
def not_enough_10():
    player = players_per_group + 1
    for week in range(3, num_weeks + 1):
        for group in range(1, num_groups + 1):
            if (group == week - 1): plus_clause([get_variable(player, group, week)])
            else: plus_clause([-get_variable(player, group, week)])

# SB11
def symmetry_breaking_11():
    diff = players_per_group - 1
    for group in range(1, num_groups + 1):
        for week in range(2, num_weeks):
            for other_week in range(week + 1, num_weeks + 1):
                plus_clause([-get_variable(week + diff, group, week), -get_variable(other_week + diff, group, other_week)])

# SB12
def symmetry_breaking_12():
    for group in range(1, num_groups + 1):
        for week in range(3, num_weeks + 1):
            for player in range(players_per_group + 1, num_players + 1):
                for other_player in range(player + players_per_group, num_players + 1, players_per_group):
                    plus_clause([-get_variable(player, group, week), -get_variable(other_player, group, week)])

def bonus_9_9_10():
    lst = [0, 2, num_groups, 1]
    for x in range (3, num_groups): lst.append(x)
    for player in range(players_per_group + 1, 2 * players_per_group + 1):
        for group in range(1, num_groups + 1):
            if (group == lst[player - players_per_group]): plus_clause([get_variable(player, group, 3)])
            else: plus_clause([-get_variable(player, group, 3)])

# SB13 + 14: Add SB10 or not_enough_10 to make it executable
def symmetry_breaking_13_14():
    for group in range(2, num_groups + 1):
        row_id = group + 1
        for week in range(3, num_weeks + 1):
            for player in range(2 * players_per_group + 1, num_players + 1, players_per_group):
                temp_id = -get_variable(player, group, week)
                for other_player in range(player + 1, player + players_per_group):
                    diff = other_player - player
                    for other_group in range(1, num_groups + 1):
                        x1 = get_variable(players_per_group + 1 + diff, other_group, row_id)
                        x2 = get_variable(other_player, other_group, week)
                        plus_clause([temp_id, x1, -x2])
                        plus_clause([temp_id, -x1, x2])

# Build the GS1 matrix
def build_gs1():
    a = {}
    for row in range(0, num_groups - 1): a[row] = []
    # row 0
    a[0] = [2, num_groups, 1]
    for x in range (3, num_groups): a[0].append(x)
    # row 1
    a[1] = [3, 1, 4]
    for i in range(3, num_groups - 1): a[1].append((a[0][i] + 2) % (num_groups + 1))
    a[1].append(2)
    # row 2
    for i in range(0, num_groups - 2): a[2].append(a[1][i] + 1)
    a[2][1] = 3
    lst = [2, 1]
    for i in lst: a[2].append(i)
    # row 3 -> last
    for row in range(3, num_groups - 1):
        lst.append(a[row - 1][1])
        for i in range(0, num_groups - row): a[row].append(a[row - 1][i] + 1)
        for i in lst: a[row].append(i)

    for row in range(0, num_groups - 1):
        week = row + 3
        for col in range(0, num_groups):
            player = players_per_group + 1 + col
            for group in range(1, num_groups + 1):
                if (group == a[row][col]): plus_clause([get_variable(player, group, week)])
                else: plus_clause([-get_variable(player, group, week)])

# SB15
def symmetry_breaking_15():
    for week in range(3, num_weeks + 1):
        player = week + players_per_group - 1
        plus_clause([get_variable(player, 1, week)])
        for group in range(2, num_groups + 1):
            plus_clause([-get_variable(player, group, week)])

# SB16
def symmetry_breaking_16():
    for gs in range (2, num_groups):
        start_player = gs * players_per_group + 1
        diff = gs * players_per_group - 1
        for player in range(start_player, start_player + players_per_group):
            for other_player in range(player + 1, start_player + players_per_group):
                for group in range(1, num_groups + 1):
                    plus_clause([-get_variable(player, group, player - diff), -get_variable(other_player, group, other_player - diff)])

# returns a unique identifier for the variable that represents the assignment of the player to the group in the week
def get_variable(player, group, week):
    player -= 1
    group -= 1
    week -= 1
    return 1 + player + (group * num_players) + (week * num_players * num_groups)

def resolve_variable(v):
    tmp = abs(v) - 1
    player = tmp % num_players + 1
    tmp //= num_players
    group = tmp % num_groups + 1
    tmp //= num_groups
    week = tmp + 1
    assert get_variable(player, group, week) == abs(v)
    return player, group, week

def validate_result(solution):
    table = {}
    for week in range(1, num_weeks + 1):
        table[week] = {}
        for group in range(1, num_groups + 1): table[week][group] = []

    for v in solution:
        if abs(v) > num_players * num_groups * num_weeks: break
        if v > 0:
            player, group, week = resolve_variable(v)
            table[week][group].append(player)

    # Check part 1
    has_played = [0 for i in range(num_players + 1)]
    for week in range(1, num_weeks + 1):
        for player in range(1, num_players + 1):
            has_played[player] = 0
        for group in range(1, num_groups + 1):
            for player in table[week][group]:
                if (has_played[player] == 1): return False
                has_played[player] = 1

    # Check part 2
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            if (len(table[week][group]) != players_per_group): return False

    # Check part 3
    play_together = [[0 for j in range(num_players + 1)] for i in range(num_players + 1)]
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            for id1 in range(0, players_per_group):
                x = table[week][group][id1]
                for id2 in range(id1 + 1, players_per_group):
                    y = table[week][group][id2]
                    if (play_together[x][y] == 1):
                        return False
                    play_together[x][y] = 1
    return True

def process_results(results):
    new_table = {}
    for week in range(1, num_weeks + 1):
        new_table[week] = {}
        for group in range(1, num_groups + 1):
            new_table[week][group] = []
    for row in results:
        new_table[row["week"]][row["group"]].append(row["player"])
    return new_table

def show_results(results):
    print_table = PrettyTable()
    field_names = ["Week"]
    for group in range(1, num_groups + 1):
        field_names.append("Group " + str(group))
    print_table.field_names = field_names
    for week in range(1, num_weeks + 1):
        row = [str(week)]
        for group in range(1, num_groups + 1):
            row.append(str(",".join(list(map(str, results[week][group])))))
        print_table.add_row(row)
    print_to_console_and_log(print_table)

def process_results2(results):
    new_table = {}
    for week in range(1, num_weeks + 1):
        new_table[week] = {}
        for player in range(1, num_players + 1):
            new_table[week][player] = []
    for row in results:
        new_table[row["week"]][row["player"]] = row["group"]
    return new_table

def show_results2(results):
    print_table = PrettyTable()
    field_names = ["W\P"]
    for player in range(1, num_players + 1):
        field_names.append(str(player))
    print_table.field_names = field_names
    for week in range(1, num_weeks + 1):
        row = [str(week)]
        for player in range(1, num_players + 1):
            row.append(str(results[week][player]))
        print_table.add_row(row)
    print_to_console_and_log(print_table)

def interrupt(s): s.interrupt()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = online_path + 'out'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try: book = load_workbook(excel_file_path)
        except BadZipFile: book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print_to_console_and_log(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")

def check_legit(solution):
    results = []
    for v in solution:
        if abs(v) > num_players * num_groups * num_weeks: break
        if v > 0 and v <= num_players * num_groups * num_weeks:
            player, group, week = resolve_variable(v)
            results.append({"player": player, "group": group, "week": week})

    final_result = process_results(results)
    show_results(final_result)

    # board = process_results2(results)
    # show_results2(board)

    print_to_console_and_log("Checking validation of the solution...")
    if (not validate_result(solution)):
        print_to_console_and_log("Invalid solution. TERMINATE right now.\n")
        return False
    else: print_to_console_and_log("Valid solution.\n")
    return True

def run_pythonsat(problem_name):
    print("Running Python SAT...")
    result_dict = {
        "ID": id_counter,
        "Problem": problem_name,
        "Type": "new_sequential",
        "SAT Solver": "Python SAT",
        "Time": "",
        "Result": "",
        "Variables": 0,
        "Clauses": 0
    }

    # Store the number of variables and clauses before solving the problem
    num_vars = sat_solver.nof_vars()
    num_clauses = sat_solver.nof_clauses()

    result_dict["Variables"] = num_vars
    result_dict["Clauses"] = num_clauses
    print_to_console_and_log("Variables: " + str(num_vars))
    print_to_console_and_log("Clauses: " + str(num_clauses))

    print_to_console_and_log("Searching for a solution...")
    timer = Timer(time_budget, interrupt, [sat_solver])
    timer.start()

    sat_status = sat_solver.solve_limited(expect_interrupt = True)

    if sat_status is False:
        elapsed_time = float(format(sat_solver.time(), ".3f"))
        print_to_console_and_log(f"UNSAT. Time run: {elapsed_time}s.\n")
        result_dict["Result"] = "unsat"
        result_dict["Time"] = elapsed_time

    else:
        solution = sat_solver.get_model()
        if solution is None:
            print_to_console_and_log(f"Time limit exceeded ({time_budget}s).\n")
            result_dict["Result"] = "timeout"
            result_dict["Time"] = time_budget
        
        else:
            elapsed_time = float(format(sat_solver.time(), ".3f"))
            print_to_console_and_log(f"A solution was found in {elapsed_time}s.")
            result_dict["Result"] = "sat"
            result_dict["Time"] = elapsed_time
            sat_accum_stats = sat_solver.accum_stats()
            print_to_console_and_log("Restarts: " + str(sat_accum_stats['restarts']) +
                    ", decisions: " + str(sat_accum_stats['decisions']) +
                    ", propagations: " + str(sat_accum_stats["propagations"]) + '\n')
            
            if not check_legit(solution):
                timer.cancel()
                sys.exit(1)

    timer.cancel()
    sat_solver.delete()
    write_to_xlsx(result_dict)

def run_kissat(problem_name):
    print("Running KiSSAT...")
    result_dict = {
        "ID": id_counter,
        "Problem": problem_name,
        "Type": "new_sequential",
        "SAT Solver": "KiSSAT",
        "Time": "",
        "Result": "",
        "Variables": 0,
        "Clauses": 0
    }

    # Store the number of variables and clauses before solving the problem
    num_vars = id_variable
    num_clauses = len(all_clauses)

    result_dict["Variables"] = num_vars
    result_dict["Clauses"] = num_clauses
    print_to_console_and_log("Variables: " + str(num_vars))
    print_to_console_and_log("Clauses: " + str(num_clauses))

    def write_to_cnf():
        # Create the directory if it doesn't exist
        input_path = online_path + "input_cnf"
        if not os.path.exists(input_path): os.makedirs(input_path)

        # Create the full path to the file "{problem}.cnf" in the directory "input_cnf"
        file_name = problem_name + ".cnf"
        file_path = os.path.join(input_path, file_name)

        # Write data to the file
        with open(file_path, 'w') as writer:
            # Write a line of information about the number of variables and constraints
            writer.write("p cnf " + str(num_vars) + " " + str(num_clauses) + "\n")

            # Write each clause to the file
            for clause in all_clauses:
                for literal in clause: writer.write(str(literal) + " ")
                writer.write("0\n")
        print_to_console_and_log("CNF written to " + file_path + ".\n")
    
    write_to_cnf()
    all_clauses.clear()

    # Create the directory if it doesn't exist
    output_path = online_path + "output_kissat"
    if not os.path.exists(output_path): os.makedirs(output_path)

    # Create the full path to the file "{problem}.txt"
    file_name = problem_name + ".txt"
    file_path = os.path.join(output_path, file_name)

    print_to_console_and_log("Searching for a solution...")
    bashCommand = f"ls input_cnf/{problem_name}.cnf | xargs -n 1 ./kissat --time={time_budget} --relaxed > {file_path}"
    os.system(bashCommand)

    def handleFile():
        result_text = "timeout"
        time_run = time_budget
        solution = []

        result = []
        with open(file_path, 'r') as file: lines = file.readlines()
        for line in lines:
            if line.strip() == "s SATISFIABLE": result_text = "sat"
            elif line.strip() == "s UNSATISFIABLE": result_text = "unsat"
            elif result_text == "sat" and line.strip().startswith("v"):
                solution.extend(map(int, line.strip().split()[1:]))
            elif result_text != "timeout" and line.strip().startswith("c process-time:"):
                tmp = line.split()
                time_run = float(tmp[len(tmp) - 2])
                break

        result_dict["Result"] = result_text
        result_dict["Time"] = time_run

        if result_text == "timeout": print_to_console_and_log(f"Time limit exceeded ({time_budget}s).\n")
        elif result_text == "sat":
            print_to_console_and_log(f"A solution was found in {time_run}s.")
            solution.pop()  # Remove the last 0 from the solution
            if not check_legit(solution): sys.exit(1)
        else: print_to_console_and_log(f"UNSAT. Time run: {time_run}s.\n")

    handleFile()
    write_to_xlsx(result_dict)

def solve_sat_problem():
    problem_name = f"{num_groups}-{players_per_group}-{num_weeks}"

    global num_players, id_variable, sat_solver, id_counter
    num_players = players_per_group * num_groups
    id_variable = num_players * num_groups * num_weeks
    id_counter += 1
    assert num_groups > 1 and players_per_group > 1

    print_to_console_and_log(
        f"Problem no. {id_counter}:\n" +
        f"Number of groups: {num_groups}.\n" +
        f"Players per group: {players_per_group}.\n" + 
        f"Number of weeks: {num_weeks}.\n")

    if enable_pythonsat: sat_solver = Glucose3(use_timer = True)
    generate_all_clauses()

    if enable_pythonsat: run_pythonsat(problem_name)
    if enable_kissat: run_kissat(problem_name)
    
    print_to_console_and_log('-' * 120)

# read input data from file data.txt (many lines, each line is a tuple of 3 integers:
# number of weeks, number of players per group, number of groups)
def run_from_input_file():
    global num_groups, players_per_group, num_weeks
    with open(online_path + 'data.txt') as f:
        for line in f:
            num_groups, players_per_group, num_weeks = map(int, line.split())
            solve_sat_problem()

    log_file.close()

if __name__ == "__main__": run_from_input_file()
