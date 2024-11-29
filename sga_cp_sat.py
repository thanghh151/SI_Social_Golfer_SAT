from ortools.sat.python import cp_model
import time
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Đọc tham số từ file data_new.txt
def read_parameters_from_file(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()
    return lines

# Đường dẫn đến file data_new.txt
file_path = 'data_new.txt'
online_path = './'  # Adjust this path as needed

# Đọc tham số từ file
lines = read_parameters_from_file(file_path)

# Open the log file in append mode
log_file = open(online_path + 'console.log', 'a')

# Define a custom print function that writes to both console and log file
def print_to_console_and_log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file=log_file, **kwargs)
    log_file.flush()

# Initialize result list
results = []

# SB1: The first week order is [1, 2, 3, ... x]
def symmetry_breaking_1(model, variables, num_players, num_groups, players_per_group, m1, m2):
    for player in range(1, num_players + 1):
        if m2 is None or len(players_per_group) == 1:
            right_group = (player - 1) // players_per_group[0] + 1
        else: 
            if player <= m2 * players_per_group[1]:
                right_group = (player - 1) // players_per_group[1] + 1
            else: 
                right_group = m2 + (player - m2 * players_per_group[1] - 1) // players_per_group[0] + 1

        # Áp dụng ràng buộc cho nhóm tương ứng
        for group in range(1, num_groups + 1):
            if group == right_group:
                model.Add(variables[(player, group, 1)] == 1)  # Người chơi ở đúng nhóm
            else:
                model.Add(variables[(player, group, 1)] == 0)  # Người chơi không ở nhóm khác



# SB2: From week 2, first p players belong to p groups
def symmetry_breaking_2(model, variables, num_players, num_groups, players_per_group, m1, m2, num_weeks):
    max_week = num_weeks // 2 - 1 if num_weeks / 2 == num_weeks // 2 else num_weeks // 2
    for week in range(2, max_week + 1):
        for player in range(1, num_groups + 1):
            for group in range(1, num_groups + 1):
                if group == player:
                    model.Add(variables[(player, group, week)] == 1)
                else:
                    model.Add(variables[(player, group, week)] == 0)


# Hàm để chạy mô hình với các tham số đầu vào
def run_model(num_players, num_weeks, players_per_group, id_counter, time_budget=600):
    # Tạo model
    model = cp_model.CpModel()

    # Tính toán số nhóm và các m1, m2
    def find_all_valid_combinations(num_players, player_per_group):
        valid_combinations = set()

        k1 = players_per_group[0]
        k2 = players_per_group[1] if len(players_per_group) > 1 else None

        # Số lượng nhóm tối đa
        max_groups = num_players // min(players_per_group)

        for num_groups in range(2, max_groups + 1):
            for m1 in range(1, num_players // k1 + 1):
                if k2 is not None:
                    for m2 in range(1, num_players // k2 + 1):
                        if m1 * k1 + m2 * k2 == num_players and m1 + m2 == num_groups:
                            valid_combinations.add((k1, k2, m1, m2, num_groups))
                else:
                    if m1 * k1 == num_players and m1 == num_groups:
                        valid_combinations.add((k1, None, m1, 0, num_groups))

        return list(valid_combinations)

    # Gọi hàm để tìm các kết hợp hợp lệ
    valid_combinations = find_all_valid_combinations(num_players, players_per_group)
    print_to_console_and_log("Valid combinations (k1, k2, m1, m2, num_groups):", valid_combinations)

    for combination in valid_combinations:
        k1, k2, m1, m2, num_groups = combination
        print_to_console_and_log(f"Running model for combination: k1={k1}, k2={k2}, m1={m1}, m2={m2}, num_groups={num_groups}")

        # Tạo biến cho mỗi người chơi, nhóm và tuần
        variables = {}
        for player in range(1, num_players + 1):
            for group in range(1, num_groups + 1):
                for week in range(1, num_weeks + 1):
                    variables[(player, group, week)] = model.NewBoolVar(f'player_{player}_group_{group}_week_{week}')

        def get_variable(golfer, group, week):
            return variables[(golfer, group, week)]

        # Ràng buộc: Áp dụng symmetry breaking
        symmetry_breaking_1(model, variables, num_players, num_groups, players_per_group, m1, m2)
        symmetry_breaking_2(model, variables, num_players, num_groups, players_per_group, m1, m2, num_weeks)

        # Ràng buộc: Mỗi người chơi chỉ xuất hiện trong một nhóm mỗi tuần
        for player in range(1, num_players + 1):
            for week in range(1, num_weeks + 1):
                model.AddExactlyOne(variables[(player, group, week)] for group in range(1, num_groups + 1))

        # Ràng buộc: Mỗi nhóm phải có một số lượng người chơi nhất định mỗi tuần
        for week in range(1, num_weeks + 1):
            if m2 > 0:
                for group in range(1, m2 + 1):
                    model.Add(sum(variables[(player, group, week)] for player in range(1, num_players + 1)) == players_per_group[1])

                for group in range(m2 + 1, num_groups + 1):
                    model.Add(sum(variables[(player, group, week)] for player in range(1, num_players + 1)) == players_per_group[0])
            else:
                for group in range(1, num_groups + 1):
                    model.Add(sum(variables[(player, group, week)] for player in range(1, num_players + 1)) == players_per_group[0])

        # Áp dụng các ràng buộc không cho phép golfer1 và golfer2 xuất hiện cùng nhau
        for week in range(1, num_weeks + 1):
            for group in range(1, num_groups + 1):
                for golfer1 in range(1, num_players + 1):
                    for golfer2 in range(golfer1 + 1, num_players + 1):
                        for other_group in range(1, num_groups + 1):
                            for other_week in range(week + 1, num_weeks + 1):
                                model.AddBoolOr([
                                    get_variable(golfer1, group, week).Not(),
                                    get_variable(golfer2, group, week).Not(),
                                    get_variable(golfer1, other_group, other_week).Not(),
                                    get_variable(golfer2, other_group, other_week).Not()
                                ])

        # Giải quyết vấn đề
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = time_budget
        start_time = time.time()
        status = solver.Solve(model)
        end_time = time.time()
        status = solver.Solve(model)

        # Hàm hiển thị giải pháp
        def display_solution():
            if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
                print_to_console_and_log("Solution Found:")
                for week in range(1, num_weeks + 1):
                    print_to_console_and_log(f"\nWeek {week}:")
                    for group in range(1, num_groups + 1):
                        group_members = []
                        for player in range(1, num_players + 1):
                            if solver.Value(variables[(player, group, week)]):
                                group_members.append(str(player))
                        print_to_console_and_log(f"  Group {group}: " + ", ".join(group_members))
            elif status == cp_model.INFEASIBLE:
                print_to_console_and_log("No solution found.")
            elif status == cp_model.MODEL_INVALID:
                print_to_console_and_log("Model invalid.")
            elif status == cp_model.UNKNOWN:
                print_to_console_and_log("Solver stopped due to time limit or other reasons.")

        # Gọi hàm hiển thị kết quả
        display_solution()

        # In ra số clause, variable, thời gian giải bài toán
        num_variables = len(model.Proto().variables)
        num_constraints = len(model.Proto().constraints)
        solving_time = solver.WallTime()
        # external_time = end_time - start_time
        print_to_console_and_log(f"Number of variables: {num_variables}")
        print_to_console_and_log(f"Number of constraints: {num_constraints}")
        print_to_console_and_log(f"Solving time: {solving_time:.3f} seconds")
        # print_to_console_and_log(f"External time: {external_time:.3f} seconds")


        # Collect results
        result_dict = {
            "ID": id_counter,
            "Problem": f"{num_players}-{players_per_group}-{num_weeks}-[{m1}-{m2}]",
            "Type": "cp-sat",
            "Time": f"{solving_time:.3f}",
            "Result": "sat" if status in [cp_model.OPTIMAL, cp_model.FEASIBLE] else "unsat" if status == cp_model.INFEASIBLE else "timeout",
            "Variables": num_variables,
            "Clauses": num_constraints
        }
        results.append(result_dict)
        write_to_xlsx(result_dict)

# Function to write results to an Excel file
def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = online_path + 'out'

    # Write the results to an Excel file
    if not os.path.exists(output_path):
        os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        book.save(excel_file_path)

    else:
        df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print_to_console_and_log(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")

# Chạy mô hình cho từng dòng trong file data_new.txt
id_counter = 1
for line in lines:
    parts = line.split()
    num_players = int(parts[0])
    players_per_group = eval(parts[1])
    num_weeks = int(parts[2])
    print_to_console_and_log(f"\nRunning model for num_players={num_players}, num_weeks={num_weeks}, players_per_group={players_per_group}")
    run_model(num_players, num_weeks, players_per_group, id_counter)
    id_counter += 1

# Close the log file
log_file.close()