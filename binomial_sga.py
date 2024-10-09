from pysat.solvers import Glucose3, Solver
from prettytable import PrettyTable
from threading import Timer
import datetime
import pandas as pd
import os
import sys
import ast
from openpyxl import load_workbook
from openpyxl import Workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

num_weeks: int  # number of weeks
players_per_group: list[int]  # players per group
num_groups: int  # number of groups
num_players: int 
id_variable: int
time_budget = 600
show_additional_info = True
online_path = ''

sat_solver: Solver

enable_kissat = False
all_clauses = []
id_counter = 0

def generate_all_clauses(m1, m2, num_groups):
    ensure_golfer_plays_at_least_once_per_week(m1, m2, num_groups)
    assign_golfers_to_groups(m1, m2, num_groups)
    ensure_golfer_plays_in_one_group_per_week(m1, m2, num_groups)
    ensure_unique_player_in_group_per_week(m1, m2, num_groups)
    ensure_unique_position_for_player_in_group(m1, m2, num_groups)
    ensure_player_in_group_if_assigned_to_week(m1, m2, num_groups)
    ensure_no_repeated_players_in_groups(m1, m2, num_groups)
    # generate_symmetry_breaking_clause1()
    # generate_symmetry_breaking_clause2()
    # generate_symmetry_breaking_clause3()

# (ALO) Every golfer plays at least once a week
# x_w_p_g (1)
def ensure_golfer_plays_at_least_once_per_week(m1, m2, num_groups):
    """
    Ensures that each golfer plays at least once per week.
    """
    for player in range(1, num_players + 1):
        for week in range(1, num_weeks + 1):
            clause = []
            for group in range(1, num_groups + 1):
                group_size = players_per_group[0] if group <= m1 else players_per_group[1]
                for position in range(1, group_size + 1):
                    clause.append(get_variable(player, position, group, week, num_groups))
            sat_solver.add_clause(clause)
            all_clauses.append(clause)

# (AMO) Each golfer plays at most once in each group each week
# x_w_p_g_p (2)
def assign_golfers_to_groups(m1, m2, num_groups):
    """
    Assigns golfers to groups for each week and position using SAT solver.
    """
    for golfer in range(1, num_players + 1):
        for week in range(1, num_weeks + 1):
            for group in range(1, num_groups + 1):
                group_size = players_per_group[0] if group <= m1 else players_per_group[1]
                for position in range(1, group_size + 1):
                    for other_position in range(position + 1, group_size + 1):
                        clause = [-1 * get_variable(golfer, position, group, week, num_groups),
                                  -1 * get_variable(golfer, other_position, group, week, num_groups)]
                        sat_solver.add_clause(clause)
                        all_clauses.append(clause)

# AMO_No golfer plays in more than one group in any week
# x_w_p_g_g_p (3)                          
def ensure_golfer_plays_in_one_group_per_week(m1, m2, num_groups):
    """
    Ensures that each golfer plays in only one group per week.
    """
    for player in range(1, num_players + 1):
        for week in range(1, num_weeks + 1):
            for group in range(1, num_groups + 1):
                group_size = players_per_group[0] if group <= m1 else players_per_group[1]
                for position in range(1, group_size + 1):
                    for next_group in range(group + 1, num_groups + 1):
                        next_group_size = players_per_group[0] if next_group <= m1 else players_per_group[1]
                        for next_position in range(1, next_group_size + 1):
                            clause = [-1 * get_variable(player, position, group, week, num_groups),
                                      -1 * get_variable(player, next_position, next_group, week, num_groups)]
                            sat_solver.add_clause(clause)
                            all_clauses.append(clause)

# (ALO) ensure each player appears only once in a group in a week
# w_g_p_x (4)                   
def ensure_unique_player_in_group_per_week(m1, m2, num_groups):
    """
    Ensures that each player appears in only one group per week.
    """
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            group_size = players_per_group[0] if group <= m1 else players_per_group[1]
            for position in range(1, group_size + 1):
                clause = []
                for golfer in range(1, num_players + 1):
                    clause.append(get_variable(golfer, position, group, week, num_groups))
                sat_solver.add_clause(clause)
                all_clauses.append(clause)

# (AMO) ensure no two players occupy the same position in the same group in the same week
# w_g_p_x_p (5)
def ensure_unique_position_for_player_in_group(m1, m2, num_groups):
    """
    Ensures that each player has a unique position within their group for each week.
    """
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            group_size = players_per_group[0] if group <= m1 else players_per_group[1]
            for position in range(1, group_size + 1):
                for golfer in range(1, num_players + 1):
                    for other_golfer in range(golfer + 1, num_players + 1):
                        clause = [-1 * get_variable(golfer, position, group, week, num_groups),
                                  -1 * get_variable(other_golfer, position, group, week, num_groups)]
                        sat_solver.add_clause(clause)
                        all_clauses.append(clause)

# This is a clause combining two sets of variables, ijkl and ikl (x_g_w_p) _6_
# ensure that if a player is in a group in a week, then they must be in one of the positions in that group, and vice versa
def ensure_player_in_group_if_assigned_to_week(m1, m2, num_groups):
    """
    Ensures that each player is assigned to a group in each week.
    """
    for golfer in range(1, num_players + 1):
        for group in range(1, num_groups + 1):
            group_size = players_per_group[0] if group <= m1 else players_per_group[1]
            for week in range(1, num_weeks + 1):
                clause = [-1 * get_variable2(golfer, group, week, num_groups)]
                for position in range(1, group_size + 1):
                    clause.append(get_variable(golfer, position, group, week, num_groups))
                    clause2 = [get_variable2(golfer, group, week, num_groups),
                               -1 * get_variable(golfer, position, group, week, num_groups)]
                    sat_solver.add_clause(clause2)
                sat_solver.add_clause(clause)
                all_clauses.append(clause)

# If two players m and n play in the same group k in week l, they cannot play together in any group together in future weeks
# w_g_x_x_g_w (7)  
def ensure_no_repeated_players_in_groups(m1, m2, num_groups):
    """
    Ensures that no players are repeated in the same group across different weeks and groups.
    """
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            for golfer1 in range(1, num_players + 1):
                for golfer2 in range(golfer1 + 1, num_players + 1):
                    for other_group in range(1, num_groups + 1):
                        for other_week in range(week + 1, num_weeks + 1):
                            clause = [-1 * get_variable2(golfer1, group, week, num_groups),
                                      -1 * get_variable2(golfer2, group, week, num_groups),
                                      -1 * get_variable2(golfer1, other_group, other_week, num_groups),
                                      -1 * get_variable2(golfer2, other_group, other_week, num_groups)]
                            sat_solver.add_clause(clause)
                            all_clauses.append(clause)

#(AMO) ensure no two players occupy the same position in the same group in the same week (x_p_g_w_x)
def generate_symmetry_breaking_clause1():
    for golfer1 in range(1, num_players + 1):
        for position1 in range(1, players_per_group):
            for group in range(1, num_groups + 1):
                for week in range(1, num_weeks + 1):
                    for golfer2 in range(1, golfer1 + 1):
                        clause = [-1 * get_variable(golfer1, position1, group, week),
                                  -1 * get_variable(golfer2, position1 + 1, group, week)]
                        sat_solver.add_clause(clause)
                        all_clauses.append(clause)

# (AMO) A player cannot be in the first position of a group in a week if they are in the first position of the next group in the same week
# x_g_w_x                        
def generate_symmetry_breaking_clause2():
    for golfer1 in range(1, num_players + 1):
        for group in range(1, num_groups):
            for week in range(1, num_weeks + 1):
                for golfer2 in range(1, golfer1):
                    clause = [-1 * get_variable(golfer1, 1, group, week),
                              -1 * get_variable(golfer2, 1, group + 1, week)]
                    sat_solver.add_clause(clause)
                    all_clauses.append(clause)

# (AMO) A player cannot be in the second position of the first group in a week if they are in the second position of the first group in the next week
def generate_symmetry_breaking_clause3():
    for golfer1 in range(1, num_players + 1):
        for week in range(1, num_weeks):
            for golfer2 in range(1, golfer1 + 1):
                clause = [-1 * get_variable(golfer1, 2, 1, week),
                          -1 * get_variable(golfer2, 2, 1, week + 1)]
                sat_solver.add_clause(clause)
                all_clauses.append(clause)


def find_all_valid_combinations():
    valid_combinations = set()
    
    k1 = players_per_group[0]
    k2 = players_per_group[1] if len(players_per_group) > 1 else None
    
    for num_groups in range(2, num_players + 1):
        for m1 in range(0, num_players + 1):
            for m2 in range(0, num_players + 1):
                if k2 is not None:
                    if k1 * m1 + k2 * m2 == num_players and m1 + m2 == num_groups:
                        valid_combinations.add((k1, k2, m1, m2, num_groups))
                else:
                    if k1 * m1 == num_players and m1 == num_groups:
                        valid_combinations.add((k1, None, m1, None, num_groups))
    
    return list(valid_combinations)

# returns a unique identifier for the variable that represents the assignment of the golfer to the position in the group in the week
def get_variable(golfer, position, group, week, num_groups):
    golfer -= 1
    position -= 1
    group -= 1
    week -= 1
    max_group_size = players_per_group[0] if len(players_per_group) == 1 else max(players_per_group[0], players_per_group[1])
    return golfer + (num_players * position) + (group * num_players * max_group_size) + (week * num_players * max_group_size * num_groups) + 1

# returns a unique identifier for the variable that represents the assignment of the golfer to the group in the week
def get_variable2(golfer, group, week, num_groups):
    golfer -= 1
    group -= 1
    week -= 1
    max_group_size = players_per_group[0] if len(players_per_group) == 1 else max(players_per_group[0], players_per_group[1])
    return golfer + (num_players * group) + (week * num_players * num_groups) + 1 + (num_players * max_group_size * num_groups * num_weeks)

def resolve_variable(v, num_groups, m1, m2):
    max_group_size = players_per_group[0] if len(players_per_group) == 1 else max(players_per_group[0], players_per_group[1])
    for golfer in range(1, num_players + 1):
        for week in range(1, num_weeks + 1):
            for group in range(1, num_groups + 1):
                group_size = players_per_group[0] if group <= m1 else players_per_group[1]
                for position in range(1, group_size + 1):
                    if abs(v) == get_variable(golfer, position, group, week, num_groups):
                        return golfer, position, group, week
    for golfer in range(1, num_players + 1):
        for week in range(1, num_weeks + 1):
            for group in range(1, num_groups + 1):
                if abs(v) == get_variable2(golfer, group, week, num_groups):
                    return golfer, group, week
    return

def validate_result(solution):
    table = {}
    for week in range(1, num_weeks + 1):
        table[week] = {}
        for group in range(1, num_groups + 1): table[week][group] = []

    for v in solution:
        if abs(v) > num_players * num_groups * num_weeks: break
        if v > 0:
            player, group, week = resolve_variable(v, num_groups, m1, m2)
            table[week][group].append(player)

    # Check part 1
    has_played = [0 for i in range(num_players + 1)]
    for week in range(1, num_weeks + 1):
        for player in range(1, num_players + 1):
            has_played[player] = 0
        for group in range(1, num_groups + 1):
            for player in table[week][group]:
                if (has_played[player] == 1): return False
                has_played[player] = 1

    # Check part 2
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            if (len(table[week][group]) != players_per_group): return False

    # Check part 3
    play_together = [[0 for j in range(num_players + 1)] for i in range(num_players + 1)]
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            for id1 in range(0, players_per_group):
                x = table[week][group][id1]
                for id2 in range(id1 + 1, players_per_group):
                    y = table[week][group][id2]
                    if (play_together[x][y] == 1):
                        return False
                    play_together[x][y] = 1
    return True

def process_results(results, num_groups):
    new_table = {}
    for week in range(1, num_weeks + 1):
        new_table[week] = {}
        for group in range(1, num_groups + 1):
            new_table[week][group] = []
    for row in results:
        new_table[row["week"]][row["group"]].append(row["golfer"])
    return new_table

def show_results(results, num_groups):
    print_table = PrettyTable()
    field_names = ["Week"]
    for group in range(1, num_groups + 1):
        field_names.append("Group " + str(group))
    print_table.field_names = field_names
    for week in range(1, num_weeks + 1):
        row = [str(week)]
        for group in range(1, num_groups + 1):
            row.append(str(",".join(list(map(str, results[week][group])))))
        print_table.add_row(row)
    print_to_console_and_log(print_table)

def process_results2(results, num_groups):
    new_table = {}
    for week in range(1, num_weeks + 1):
        new_table[week] = {}
        for player in range(1, num_players + 1):
            new_table[week][player] = []
    for row in results:
        new_table[row["week"]][row["golfer"]] = row["group"]
    return new_table

def show_results2(results, num_groups):
    print_table = PrettyTable()
    field_names = ["W\P"]
    for player in range(1, num_players + 1):
        field_names.append(str(player))
    print_table.field_names = field_names
    for week in range(1, num_weeks + 1):
        row = [str(week)]
        for player in range(1, num_players + 1):
            row.append(str(results[week][player]))
        print_table.add_row(row)
    print_to_console_and_log(print_table)

def interrupt(s): s.interrupt()

def write_to_cnf(num_vars, num_clauses, problem_name):
    # Create the directory if it doesn't exist
    input_path = online_path + "input_cnf"
    if not os.path.exists(input_path): os.makedirs(input_path)

    # Create the full path to the file "{problem}.cnf" in the directory "input_cnf"
    file_name = problem_name + ".cnf"
    file_path = os.path.join(input_path, file_name)

    # Write data to the file
    with open(file_path, 'w') as writer:
        # Write a line of information about the number of variables and constraints
        writer.write("p cnf " + str(num_vars) + " " + str(num_clauses) + "\n")

        # Write each clause to the file
        for clause in all_clauses:
            for literal in clause: writer.write(str(literal) + " ")
            writer.write("0\n")
    
    print_to_console_and_log("CNF written to " + file_path + ".\n")

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = online_path + 'out'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print_to_console_and_log(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")

def check_legit(solution, num_groups, m1, m2):
    results = []
    for v in solution:
        if v > 0:
            ikl = resolve_variable(v, num_groups, m1, m2)
            if len(ikl) == 3:
                golfer, group, week = ikl
                results.append({"golfer": golfer, "group": group, "week": week})

    final_result = process_results(results, num_groups)
    show_results(final_result, num_groups)

    board = process_results2(results, num_groups)
    show_results2(board, num_groups)

    # print_to_console_and_log("Checking validation of the solution...")
    # if (not validate_result(solution)):
    #     print_to_console_and_log("Invalid solution. TERMINATE right now.\n")
    #     return False
    # else: print_to_console_and_log("Valid solution.\n")
    return True

def run_kissat(problem_name):
    # Create the directory if it doesn't exist
    input_path = online_path + "output_kissat"
    if not os.path.exists(input_path): os.makedirs(input_path)

    # Create the full path to the file "{problem}.txt"
    file_name = problem_name + ".txt"
    file_path = os.path.join(input_path, file_name)

    print_to_console_and_log("Running KiSSAT...")
    bashCommand = f"ls input_cnf/{problem_name}.cnf | xargs -n 1 ./kissat --time={time_budget} --relaxed > output_kissat/{problem_name}.txt"
    os.system(bashCommand)
    print_to_console_and_log("KiSSAT finished.")

# solve the problem using the SAT Solver and write the results to xlsx file
def solve_sat_problem():
    global num_players, id_variable, sat_solver, id_counter

    
    valid_combinations = find_all_valid_combinations()
    for k1, k2, m1, m2, num_groups in valid_combinations:
        id_variable = num_players * num_groups * num_weeks
        id_counter += 1
        # print(f"Trying k1 = {k1}, k2 = {k2}, m1 = {m1}, m2 = {m2}, num_groups = {num_groups}...")
        result_dict = {
            "ID": id_counter,
            "Problem": f"{num_players}-{num_groups}-{players_per_group}-{num_weeks}",
            "Type": "sga",
            "Time": "",
            "Result": "",
            "Variables": 0,
            "Clauses": 0
        }

        print_to_console_and_log(
            f"Problem no. {id_counter}:\n" +
            f"Number of players: {num_players}.\n" +
            f"Number of groups: {num_groups}.\n" +
            f"Players per group: {players_per_group}.\n" +
            f"k1: {k1}.\n" +
            f"k2: {k2}.\n" +
            f"m1: {m1}.\n" + 
            f"m2: {m2}.\n" +
            f"Number of weeks: {num_weeks}.\n")

        assert num_groups > 1 and players_per_group[0] > 1

        sat_solver = Glucose3(use_timer = True)
        generate_all_clauses(m1, m2, num_groups)
        # print(sat_solver.get_model(), sat_solver.nof_clauses(), sat_solver.nof_vars())

        # Store the number of variables and clauses before solving the problem
        problem_name = f"{num_players}-{num_groups}-{players_per_group}-{num_weeks}"
        if not enable_kissat:
            num_vars = sat_solver.nof_vars()
            num_clauses = sat_solver.nof_clauses()
        else:
            num_vars = id_variable
            assert num_vars == sat_solver.nof_vars()
            num_clauses = len(all_clauses)
            # print_to_console_and_log(f"{num_clauses} {sat_solver.nof_clauses()}")

        result_dict["Variables"] = num_vars
        result_dict["Clauses"] = num_clauses
        if show_additional_info:
            print_to_console_and_log("Variables: " + str(num_vars))
            print_to_console_and_log("Clauses: " + str(num_clauses))

        print_to_console_and_log("Searching for a solution...")
        timer = Timer(time_budget, interrupt, [sat_solver])
        timer.start()

        sat_status = sat_solver.solve_limited(expect_interrupt = True)
        # print(sat_solver.get_model(), sat_solver.nof_clauses(), sat_solver.nof_vars())
        if sat_status is False:
            elapsed_time = format(sat_solver.time(), ".3f")
            print_to_console_and_log(f"UNSAT. Time run: {elapsed_time}s.\n")
            result_dict["Result"] = "unsat"
            result_dict["Time"] = elapsed_time
            # print(sat_solver.get_model(), sat_solver.nof_clauses(), sat_solver.nof_vars())

        else:
            solution = sat_solver.get_model()
            # print(sat_solver.get_model(), sat_solver.nof_clauses(), sat_solver.nof_vars())
            if solution is None:
                print_to_console_and_log(f"Time limit exceeded ({time_budget}s).\n")
                result_dict["Result"] = "timeout"
                result_dict["Time"] = time_budget
            
            else:
                elapsed_time = format(sat_solver.time(), ".3f")
                print_to_console_and_log(f"A solution was found in {elapsed_time}s.")
                result_dict["Result"] = "sat"
                result_dict["Time"] = elapsed_time

                if show_additional_info:
                    sat_accum_stats = sat_solver.accum_stats()
                    print_to_console_and_log("Restarts: " + str(sat_accum_stats['restarts']) +
                            ", decisions: " + str(sat_accum_stats['decisions']) +
                            ", propagations: " + str(sat_accum_stats["propagations"]) + '\n')
                if not check_legit(solution, num_groups, m1, m2):
                    timer.cancel()
                    sys.exit(1)

        timer.cancel()
        sat_solver.delete()
        
        if enable_kissat:
            write_to_cnf(num_vars, num_clauses, problem_name)
            run_kissat(problem_name)
        write_to_xlsx(result_dict)
        all_clauses.clear()
        
        print_to_console_and_log('-' * 120)

# Open the log file in append mode
log_file = open(online_path + 'console.log', 'a')

# Define a custom print function that writes to both console and log file
def print_to_console_and_log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file = log_file, **kwargs)
    log_file.flush()

# read input data from file data.txt (many lines, each line is number of weeks, number of players per group, number of groups)
# solve the problem
def run_from_input_file():
    global num_players, players_per_group, num_weeks
    with open('data_new.txt') as f:
        for line in f:
            parts = line.split()
            try:
                num_players = int(parts[0])
                
                # Clean and validate the list part
                list_part = parts[1].strip()
                if not (list_part.startswith('[') and list_part.endswith(']')):
                    raise SyntaxError("List part is not properly formatted")
                
                players_per_group = ast.literal_eval(list_part)
                num_weeks = int(parts[2])
                solve_sat_problem()
            except (ValueError, SyntaxError) as e:
                print(f"Error parsing line: {line.strip()}")
                print(f"Exception: {e}")
                continue

    log_file.close()


if __name__ == "__main__": run_from_input_file()
