from pysat.solvers import Glucose3, Solver
from prettytable import PrettyTable
from threading import Timer
import datetime
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import math
from datetime import datetime

num_weeks: int  # number of weeks
players_per_group: int  # players per group
num_groups: int  # number of groups
num_players: int  # players per group * number of groups
time_budget = 600
show_additional_info = True
show_additional_info_str = "Yes"

sat_solver: Solver

all_clauses = []
id_counter = 1

def generate_all_clauses():
    ensure_golfer_plays_at_least_once_per_week()
    assign_golfers_to_groups()
    ensure_golfer_plays_in_one_group_per_week()
    ensure_unique_player_in_group_per_week()
    ensure_unique_position_for_player_in_group()
    ensure_player_in_group_if_assigned_to_week()
    ensure_no_repeated_players_in_groups()
    generate_symmetry_breaking_clause1()
    generate_symmetry_breaking_clause2()
    generate_symmetry_breaking_clause3()

# (ALO) Every golfer plays at least once a week
def ensure_golfer_plays_at_least_once_per_week():
    for player in range(1, num_players + 1):
        for week in range(1, num_weeks + 1):
            clause = []
            for position in range(1, players_per_group + 1):
                for group in range(1, num_groups + 1):
                    clause.append(get_variable(player, position, group, week))
            print(clause)
            sat_solver.add_clause(clause)
            all_clauses.append(clause)


# (AMO) Each golfer plays at most once in each group each week
def assign_golfers_to_groups():
    """
    Assigns golfers to groups for each week and position using SAT solver.
    """
    k = len(bin(num_players)[2:])  # number of binary digits needed to represent num_players
    for golfer in range(1, num_players + 1):
        golfer_binary = [int(x) for x in bin(golfer)[2:].zfill(k)]  # binary representation of golfer
        for week in range(1, num_weeks + 1):
            for position in range(1, players_per_group + 1):
                for group in range(1, num_groups + 1):
                    Y = [0] * k  # new variables Y1, Y2, ..., Yk
                    for j in range(k):
                        if golfer_binary[j] == 1:
                            Y[j] = get_variable(golfer, position, group, week)  # assign Yj to Xi if j-th digit is 1
                        else:
                            Y[j] = -1 * get_variable(golfer, position, group, week)  # assign ¬Yj to Xi if j-th digit is 0
                    for other_position in range(position + 1, players_per_group + 1):
                        clause = [-1 * get_variable(golfer, position, group, week),
                                  -1 * get_variable(golfer, other_position, group, week)] + Y
                        sat_solver.add_clause(clause)
                        all_clauses.append(clause)


# AMO_No golfer plays in more than one group each week
def ensure_golfer_plays_in_one_group_per_week():
    k = len(bin(num_players)[2:])  # number of binary digits needed to represent num_players
    for player in range(1, num_players + 1):
        player_binary = [int(x) for x in bin(player)[2:].zfill(k)]  # binary representation of player
        for week in range(1, num_weeks + 1):
            for group in range(1, num_groups + 1):
                Y = [0] * k  # new variables Y1, Y2, ..., Yk
                for position in range(1, players_per_group + 1):
                    for j in range(k):
                        if player_binary[j] == 1:
                            Y[j] = get_variable(player, position, group, week)  # assign Yj to Xi if j-th digit is 1
                        else:
                            Y[j] = -1 * get_variable(player, position, group, week)  # assign ¬Yj to Xi if j-th digit is 0
                    clause = [-1 * get_variable(player, position, group, week)] + Y
                    sat_solver.add_clause(clause)
                    all_clauses.append(clause)

# (ALO) ensure each player appears only once in a group in a week
def ensure_unique_player_in_group_per_week():
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            for position in range(1, players_per_group + 1):
                clause = []
                for golfer in range(1, num_players + 1):
                    clause.append(get_variable(golfer, position, group, week))
                sat_solver.add_clause(clause)
                all_clauses.append(clause)

# (ALO) ensure no two players occupy the same position in the same group in the same week
def ensure_unique_position_for_player_in_group():
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            for position in range(1, players_per_group + 1):
                for golfer in range(1, num_players + 1):
                    for other_position in range(position + 1, players_per_group + 1):
                        clause = [-1 * get_variable(golfer, position, group, week),
                                  -1 * get_variable(golfer, other_position, group, week)]
                        sat_solver.add_clause(clause)
                        all_clauses.append(clause)


# This is a clause combining two sets of variables, ijkl and ikl
# ensure that if a player is in a group in a week, then they must be in one of the positions in that group, and vice versa
def ensure_player_in_group_if_assigned_to_week():
    for golfer in range(1, num_players + 1):
        for group in range(1, num_groups + 1):
            for week in range(1, num_weeks + 1):
                clause = [-1 * get_variable2(golfer, group, week)]
                for position in range(1, players_per_group + 1):
                    clause.append(get_variable(golfer, position, group, week))
                    clause2 = [get_variable2(golfer, group, week),
                               -1 * get_variable(golfer, position, group, week)]
                    sat_solver.add_clause(clause2)
                sat_solver.add_clause(clause)
                all_clauses.append(clause)


# If two players m and n play in the same group k in week l, they cannot play together in any group together in future weeks
def ensure_no_repeated_players_in_groups():
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            for golfer1 in range(1, num_players + 1):
                for golfer2 in range(golfer1 + 1, num_players + 1):
                    for other_group in range(1, num_groups + 1):
                        for other_week in range(week + 1, num_weeks + 1):
                            clause = [-1 * get_variable2(golfer1, group, week),
                                      -1 * get_variable2(golfer2, group, week),
                                      -1 * get_variable2(golfer1, other_group, other_week),
                                      -1 * get_variable2(golfer2, other_group, other_week)]
                            sat_solver.add_clause(clause)
                            all_clauses.append(clause)

#(ALO) ensure no two players occupy the same position in the same group in the same week
def generate_symmetry_breaking_clause1():
    for golfer1 in range(1, num_players + 1):
        for position1 in range(1, players_per_group):
            for group in range(1, num_groups + 1):
                for week in range(1, num_weeks + 1):
                    for golfer2 in range(1, golfer1 + 1):
                        clause = [-1 * get_variable(golfer1, position1, group, week),
                                  -1 * get_variable(golfer2, position1 + 1, group, week)]
                        sat_solver.add_clause(clause)
                        all_clauses.append(clause)

# A player cannot be in the first position of a group in a week if they are in the first position of the next group in the same week
def generate_symmetry_breaking_clause2():
    for golfer1 in range(1, num_players + 1):
        for group in range(1, num_groups):
            for week in range(1, num_weeks + 1):
                for golfer2 in range(1, golfer1):
                    clause = [-1 * get_variable(golfer1, 1, group, week),
                              -1 * get_variable(golfer2, 1, group + 1, week)]
                    sat_solver.add_clause(clause)
                    all_clauses.append(clause)

#A player cannot be in the second position of the first group in a week if they are in the second position of the first group in the next week
def generate_symmetry_breaking_clause3():
    for golfer1 in range(1, num_players + 1):
        for week in range(1, num_weeks):
            for golfer2 in range(1, golfer1 + 1):
                clause = [-1 * get_variable(golfer1, 2, 1, week),
                          -1 * get_variable(golfer2, 2, 1, week + 1)]
                sat_solver.add_clause(clause)
                all_clauses.append(clause)

# returns a unique identifier for the variable that represents the assignment of the golfer to the position in the group in the week
def get_variable(golfer, position, group, week):
    golfer -= 1
    position -= 1
    group -= 1
    week -= 1
    return golfer + (num_players * position) + (group * num_players * players_per_group) + (week * num_players * players_per_group * num_groups) + 1

# returns a unique identifier for the variable that represents the assignment of the golfer to the group in the week
def get_variable2(golfer, group, week):
    golfer -= 1
    group -= 1
    week -= 1
    return golfer + (num_players * group) + (week * num_players * num_groups) + 1 + (num_players * players_per_group * num_groups * num_weeks)

def resolve_variable(v):
    for golfer in range(1, num_players + 1):
        for week in range(1, num_weeks + 1):
            for position in range(1, players_per_group + 1):
                for group in range(1, num_groups + 1):
                    if abs(v) == get_variable(golfer, position, group, week):
                        return golfer, position, group, week
    for golfer in range(1, num_players + 1):
        for week in range(1, num_weeks + 1):
            for group in range(1, num_groups + 1):
                if abs(v) == get_variable2(golfer, group, week):
                    return golfer, group, week
    return

def process_results(results):
    new_table = {}
    for week in range(1, num_weeks + 1):
        new_table[week] = {}
        for group in range(1, num_groups + 1):
            new_table[week][group] = []
    for row in results:
        new_table[row["week"]][row["group"]].append(row["golfer"])
    return new_table

def show_results(results):
    print_table = PrettyTable()
    field_names = ["Week"]
    for group in range(1, num_groups + 1):
        field_names.append("Group " + str(group))
    print_table.field_names = field_names
    for week in range(1, num_weeks + 1):
        row = [str(week)]
        for group in range(1, num_groups + 1):
            row.append(str(",".join(list(map(str, results[week][group])))))
        print_table.add_row(row)
    print(print_table)

def change_time_budget():
    global time_budget
    while True:
        try:
            time_budget = int(input("\nEnter new time limit in seconds (current: " + str(time_budget) + "s): "))
            if time_budget < 0:
                time_budget = 0
            elif time_budget > 999999:
                time_budget = 999999
        except ValueError:
            print("Enter a valid value\n")
            continue
        else:
            break

def change_showing_additional_info():
    global show_additional_info, show_additional_info_str
    while True:
        try:
            print(
                "\nShould additional information about solving the problem in the SAT Solver be displayed (i.e.: number of variables, number of clauses, propagations, conflicts, decisions and restarts)")
            print("1 - Yes")
            print("2 - No")
            choice = int(input("Choose an option: "))
            if choice == 1:
                show_additional_info = True
                show_additional_info_str = "Yes"
            elif choice == 2:
                show_additional_info = False
                show_additional_info_str = "No"
            else:
                print("Enter a valid value\n")
                continue
        except ValueError:
            print("Enter a valid value\n")
            continue
        else:
            break

# def main_menu():
#     while True:
#         try:
#             print("\n/------------------------------\\")
#             print("| Social Golfer Problem Solver |")
#             print("\\------------------------------/")
#             print("1 - Solve the Social Golfer problem")
#             print("2 - Change timeout (currently: " + str(time_budget) + "s)")
#             print("3 - Change showing additional information (currently: " + show_additional_info_str + ")")
#             print("0 - Finish")
#             selection = int(input("Select options: "))
#             if selection == 1:
#                 menu()
#             elif selection == 2:
#                 change_time_budget()
#             elif selection == 3:
#                 change_showing_additional_info()
#             elif selection == 0:
#                 return

#         except ValueError:
#             print("Enter a valid value\n")
#             continue
#         else:
#             pass

# def menu():
#     global num_weeks, players_per_group, num_groups
#     while True:
#         try:
#             num_weeks = int(input("Enter the number of weeks: "))
#             if num_weeks <= 0:
#                 print("Enter a valid value\n")
#                 continue
#             players_per_group = int(input("Enter the number of players per group: "))
#             if players_per_group <= 0:
#                 print("Enter a valid value\n")
#                 continue
#             num_groups = int(input("Enter the number of groups: "))
#             if num_groups <= 0:
#                 print("Enter a valid value\n")
#                 continue
#         except ValueError:
#             print("Enter a valid value\n")
#             continue
#         else:
#             break
#     solve_sat_problem()

def interrupt(s):
    s.interrupt()

# solve the problem using the SAT Solver and write the results to xlsx file
def solve_sat_problem():
    global num_players, sat_solver
    num_players = players_per_group * num_groups
    
    # Clear the all_clauses list
    all_clauses.clear()

    print("\nGenerating a problem.")

    sat_solver = Glucose3(use_timer=True)
    generate_all_clauses()
    
    # Store the number of variables and clauses before solving the problem
    num_vars = sat_solver.nof_vars()
    num_clauses = sat_solver.nof_clauses()

    if show_additional_info:
        print("Clauses: " + str(sat_solver.nof_clauses()))
        print("Variables: " + str(sat_solver.nof_vars()))

    print("\nSearching for a solution.")

    timer = Timer(time_budget, interrupt, [sat_solver])
    timer.start()

    start_time = time.time()
    sat_status = sat_solver.solve_limited(expect_interrupt=True)

    global id_counter

    result_dict = {
        "ID": id_counter,
        "Problem": f"{num_weeks}-{players_per_group}-{num_groups}",
        "Type": "binary",
        "Time": "",
        "Result": "",
        "Variables": 0,
        "Clauses": 0
    }
    
    id_counter += 1
    if sat_status is False:
        end_time = time.time()
        elapsed_time = end_time - start_time
        print("Not found. Time exceeded (" + '{0:.3f}s'.format(elapsed_time) + ").\n")
        result_dict["Result"] = "unsat"
        result_dict["Time"] = '{0:.3f}'.format(elapsed_time)
        result_dict["Variables"] = sat_solver.nof_vars()
        result_dict["Clauses"] = sat_solver.nof_clauses()
    else:
        solution = sat_solver.get_model()
        if solution is None:
            end_time = time.time()
            elapsed_time = end_time - start_time
            if elapsed_time > 10:
                print("Timeout exceeded (" + '{0:.3f}s'.format(elapsed_time) + ").\n")
                result_dict["Result"] = "timeout"
            else:
                print("Not found. Time exceeded (" + '{0:.3f}s'.format(elapsed_time) + ").\n")
                result_dict["Result"] = "unsat"
            result_dict["Time"] = '{0:.3f}'.format(elapsed_time)
            result_dict["Variables"] = sat_solver.nof_vars()
            result_dict["Clauses"] = sat_solver.nof_clauses()
        else:
            print(
                "A solution was found in time " + '{0:.3f}s'.format(sat_solver.time()) + ". Generating it now.\n")
            result_dict["Result"] = "sat"

            results = []
            for v in solution:
                if v > 0:
                    ijkl = resolve_variable(v)
                    if len(ijkl) == 3:
                        golfer, group, week = ijkl
                        results.append({"golfer": golfer, "group": group, "week": week})

            final_result = process_results(results)
            show_results(final_result)

            if show_additional_info:
                sat_accum_stats = sat_solver.accum_stats()
                print("Restarts: " +
                        str(sat_accum_stats['restarts']) +
                        ", conflicts: " +
                        ", decisions: " +
                        str(sat_accum_stats['decisions']) +
                        ", propagations: " +
                        str(sat_accum_stats["propagations"]))

            result_dict["Time"] = '{0:.3f}'.format(sat_solver.time())
            result_dict["Variables"] = sat_solver.nof_vars()
            result_dict["Clauses"] = sat_solver.nof_clauses()

            sat_solver.delete()

    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    # Write the results to an Excel file
    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"out/results_{current_date}.xlsx"
        
    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']

        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        book.save(excel_file_path)

    else:
        df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print("Result written to Excel file:", os.path.abspath(excel_file_path))  # Print full path
    print("Result added to Excel file.")
    
    # Create the directory if it doesn't exist
    directory_path = "input_v1"
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)

    # Create the full path to the file "{problem}.cnf" in the directory "input_v1"
    problem_name = f"{num_weeks}-{players_per_group}-{num_groups}-binary"
    file_name = problem_name + ".cnf"
    file_path = os.path.join(directory_path, file_name)

    # Write data to the file
    with open(file_path, 'w') as writer:
        # Write a line of information about the number of variables and constraints
        writer.write("p cnf " + str(num_vars) + " " + str(num_clauses) + "\n")

        # Write each clause to the file
        for clause in all_clauses:
            for literal in clause:
                writer.write(str(literal) + " ")
            writer.write("0\n")

    print("CNF written to " + file_path)


# read input data from file data.txt (many lines, each line is number of weeks, number of players per group, number of groups)
# solve the problem

def run_from_input_file():
    global num_weeks, players_per_group, num_groups
    with open("data.txt") as f:
        for line in f:
            num_weeks, players_per_group, num_groups = map(int, line.split())
            solve_sat_problem()

if __name__ == "__main__":
    # main_menu()
    run_from_input_file()