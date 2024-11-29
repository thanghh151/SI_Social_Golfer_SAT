from pysat.solvers import Glucose3, Solver
from prettytable import PrettyTable
from threading import Timer
import datetime
import pandas as pd
import os
import sys
import ast
import math
from openpyxl import load_workbook
from openpyxl import Workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

num_weeks: int  # number of weeks
players_per_group: list[int]  # players per group
num_groups: int  # number of groups
num_players: int
id_variable: int
time_budget = 600
show_additional_info = True
online_path = ''

sat_solver: Solver

enable_kissat = False
all_clauses = []
id_counter = 0

def generate_all_clauses(m1, m2, num_groups):
    symmetry_breaking_1(m1, m2, num_groups)
    symmetry_breaking_2(m1, m2, num_groups)
    ensure_golfer_plays_exactly_once_per_week(num_groups)
    ensure_group_contains_exactly_p_players(m1, m2, num_groups)
    ensure_no_repeated_players_in_groups(num_groups)

def plus_clause(clause):
    sat_solver.add_clause(clause)
    if (enable_kissat): all_clauses.append(clause)

# (EO) Using binomial
def exactly_one(var: list[int], num_groups):
    n = len(var)
    assert n == num_groups

    # (1): (ALO)
    clause = []
    for i in range(0, n):
        clause.append(var[i])
    plus_clause(clause)

    # (2): (AMO)
    for i in range (0, n):
        for j in range (i + 1, n):
            plus_clause([-1 * var[i], -1 * var[j]])

# Every player plays exactly once a week
# x_w_g (1)
def ensure_golfer_plays_exactly_once_per_week(num_groups):
    for player in range(1, num_players + 1):
        for week in range(1, num_weeks + 1):
            list = []
            for group in range(1, num_groups + 1):
                list.append(get_variable(player, group, week, num_groups))
            exactly_one(list, num_groups)

# (EK) Using New Sequential encounter (NSC)
def exactly_k(var: list[int], k):
    global id_variable
    n = len(var) - 1
    assert n == num_players
    map_register = [[0 for j in range(k + 1)] for i in range(n)]
    for i in range(1, n):
        for j in range(1, min(i, k) + 1):
            id_variable += 1
            map_register[i][j] = id_variable

    # (1): If a bit is true, the first bit of the corresponding register is true
    for i in range(1, n):
        plus_clause([-1 * var[i], map_register[i][1]])

    # (2): R[i - 1][j] = 1, R[i][j] = 1;
    for i in range(2, n):
        for j in range(1, min(i - 1, k) + 1):
            plus_clause([-1 * map_register[i - 1][j], map_register[i][j]])

    # (3): If bit i is on and R[i - 1][j - 1] = 1, R[i][j] = 1;
    for i in range(2, n):
        for j in range(2, min(i, k) + 1):
            plus_clause([-1 * var[i], -1 * map_register[i - 1][j - 1], map_register[i][j]])

    # (4): If bit i is off and R[i - 1][j] = 0, R[i][j] = 0;
    for i in range(2, n):
        for j in range(1, min(i - 1, k) + 1):
            plus_clause([var[i], map_register[i - 1][j], -1 * map_register[i][j]])

    # (5): If bit i is off, R[i][i] = 0;
    for i in range(1, k + 1):
        plus_clause([var[i], -1 * map_register[i][i]])

    # (6): If R[i - 1][j - 1] = 0, R[i][j] = 0;
    for i in range(2, n):
        for j in range(2, min(i, k) + 1):
            plus_clause([map_register[i - 1][j - 1], -1 * map_register[i][j]])

    # (7): (At least k) R[n - 1][k] = 1 or (n-th bit is true and R[n - 1][k - 1] = 1)
    plus_clause([map_register[n - 1][k], var[n]])
    plus_clause([map_register[n - 1][k], map_register[n - 1][k - 1]])
    # plus_clause([map_register[n - 1][k - 1]])

    # (8): (At most k) If i-th bit is true, R[i - 1][k] = 0;
    for i in range(k + 1, n + 1):
        plus_clause([-1 * var[i], -1 * map_register[i - 1][k]])

# A group contains exactly p players
# w_g_x (2)
def ensure_group_contains_exactly_p_players(m1, m2, num_groups):
    for week in range(1, num_weeks + 1):
        if m2 is not None:
            # Groups from 1 to m2 with players_per_group[1] members
            for group in range(1, m2 + 1):
                list_vars = [-1]
                for player in range(1, num_players + 1):
                    list_vars.append(get_variable(player, group, week, num_groups))
                exactly_k(list_vars, players_per_group[1])
            # Groups from m2 + 1 to num_groups with players_per_group[0] members
            for group in range(m2 + 1, num_groups + 1):
                list_vars = [-1]
                for player in range(1, num_players + 1):
                    list_vars.append(get_variable(player, group, week, num_groups))
                exactly_k(list_vars, players_per_group[0])
        else:
            # All groups with players_per_group[0] members
            for group in range(1, num_groups + 1):
                list_vars = [-1]
                for player in range(1, num_players + 1):
                    list_vars.append(get_variable(player, group, week, num_groups))
                exactly_k(list_vars, players_per_group[0])

# Ensures that no players are repeated in the same group across different weeks and groups.
# w_g_x_x_g_w (3)
def ensure_no_repeated_players_in_groups(num_groups):
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            for golfer1 in range(1, num_players + 1):
                for golfer2 in range(golfer1 + 1, num_players + 1):
                    for other_group in range(1, num_groups + 1):
                        for other_week in range(week + 1, num_weeks + 1):
                            clause = [-1 * get_variable(golfer1, group, week, num_groups),
                                      -1 * get_variable(golfer2, group, week, num_groups),
                                      -1 * get_variable(golfer1, other_group, other_week, num_groups),
                                      -1 * get_variable(golfer2, other_group, other_week, num_groups)]
                            plus_clause(clause)
# SB1: The first week order is [1, 2, 3, ... x]
def symmetry_breaking_1(m1, m2, num_groups):
    for player in range(1, num_players + 1):
        if m2 is None:
            if player <= m1 * players_per_group[0]:
                right_group = (player - 1) // players_per_group[0] + 1
            else:
                right_group = m1 + (player - m1 * players_per_group[0] - 1) // players_per_group[1] + 1
        else:
            if player <= m2 * players_per_group[1]:
                right_group = (player - 1) // players_per_group[1] + 1
            else:
                right_group = m2 + (player - m2 * players_per_group[1] - 1) // players_per_group[0] + 1

        for group in range(1, num_groups + 1):
            if group == right_group:
                sat_solver.add_clause([get_variable(player, group, 1, num_groups)])
            else:
                sat_solver.add_clause([-1 * get_variable(player, group, 1, num_groups)])

# SB2: From week 2, first p players belong to p groups
def symmetry_breaking_2(m1, m2, num_groups):
    # max_week = math.floor((num_players - players_per_group[1]) / players_per_group[0]) - 1
    max_week = num_weeks // 2 - 1 if num_weeks / 2 == num_weeks // 2 else num_weeks // 2
    if m2 is not None:
        for week in range(2, max_week + 1):
            for player in range(1, min(num_groups, players_per_group[1]) + 1):
                for group in range(1, num_groups + 1):
                    if group == player:
                        plus_clause([get_variable(player, group, week, num_groups)])
                    else:
                        plus_clause([-1 * get_variable(player, group, week, num_groups)])
    else:
        for week in range(2, max_week + 1):
            for player in range(1, min(num_groups, players_per_group[0]) + 1):
                for group in range(1, num_groups + 1):
                    if group == player:
                        plus_clause([get_variable(player, group, week, num_groups)])
                    else:
                        plus_clause([-1 * get_variable(player, group, week, num_groups)])

# def find_valid_m1_m2():
#     k1 = players_per_group[0]
#     k2 = players_per_group[1] if len(players_per_group) > 1 else None

#     valid_combinations = []
#     for m1 in range(num_groups + 1):
#         remaining_players = num_players - m1 * k1
#         if k2 and remaining_players % k2 == 0:
#             m2 = remaining_players // k2
#             if m1 + m2 == num_groups:
#                 valid_combinations.append((m1, m2))
#         elif not k2 and remaining_players == 0 and m1 == num_groups:
#             valid_combinations.append((m1, 0))

#     return valid_combinations

def find_all_valid_combinations():
    valid_combinations = set()

    k1 = players_per_group[0]
    k2 = players_per_group[1] if len(players_per_group) > 1 else None

    for num_groups in range(2, num_players + 1):
        for m1 in range(1, num_players + 1):
            for m2 in range(1, num_players + 1):
                if k2 is not None:
                    if k1 * m1 + k2 * m2 == num_players and m1 + m2 == num_groups:
                        valid_combinations.add((k1, k2, m1, m2, num_groups))
                else:
                    if k1 * m1 == num_players and m1 == num_groups:
                        valid_combinations.add((k1, None, m1, None, num_groups))

    return list(valid_combinations)

# returns a unique identifier for the variable that represents the assignment of the player to the group in the week
def get_variable(player, group, week, num_groups):
    player -= 1
    group -= 1
    week -= 1
    return 1 + player + (group * num_players) + (week * num_players * num_groups)

def resolve_variable(v, num_groups):
    tmp = abs(v) - 1
    player = tmp % num_players + 1
    tmp //= num_players
    group = tmp % num_groups + 1
    tmp //= num_groups
    week = tmp + 1
    assert get_variable(player, group, week, num_groups) == abs(v)
    return player, group, week

def validate_result(solution):
    table = {}
    for week in range(1, num_weeks + 1):
        table[week] = {}
        for group in range(1, num_groups + 1): table[week][group] = []

    for v in solution:
        if abs(v) > num_players * num_groups * num_weeks: break
        if v > 0:
            player, group, week = resolve_variable(v, num_groups)
            table[week][group].append(player)

    # Check part 1
    has_played = [0 for i in range(num_players + 1)]
    for week in range(1, num_weeks + 1):
        for player in range(1, num_players + 1):
            has_played[player] = 0
        for group in range(1, num_groups + 1):
            for player in table[week][group]:
                if (has_played[player] == 1): return False
                has_played[player] = 1

    # Check part 2
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            if (len(table[week][group]) != players_per_group): return False

    # Check part 3
    play_together = [[0 for j in range(num_players + 1)] for i in range(num_players + 1)]
    for week in range(1, num_weeks + 1):
        for group in range(1, num_groups + 1):
            for id1 in range(0, players_per_group):
                x = table[week][group][id1]
                for id2 in range(id1 + 1, players_per_group):
                    y = table[week][group][id2]
                    if (play_together[x][y] == 1):
                        return False
                    play_together[x][y] = 1
    return True

def process_results(results, num_groups):
    new_table = {}
    for week in range(1, num_weeks + 1):
        new_table[week] = {}
        for group in range(1, num_groups + 1):
            new_table[week][group] = []
    for row in results:
        new_table[row["week"]][row["group"]].append(row["player"])
    return new_table

def show_results(results, num_groups):
    print_table = PrettyTable()
    field_names = ["Week"]
    for group in range(1, num_groups + 1):
        field_names.append("Group " + str(group))
    print_table.field_names = field_names
    for week in range(1, num_weeks + 1):
        row = [str(week)]
        for group in range(1, num_groups + 1):
            row.append(str(",".join(list(map(str, results[week][group])))))
        print_table.add_row(row)
    print_to_console_and_log(print_table)

def process_results2(results, num_groups):
    new_table = {}
    for week in range(1, num_weeks + 1):
        new_table[week] = {}
        for player in range(1, num_players + 1):
            new_table[week][player] = []
    for row in results:
        new_table[row["week"]][row["player"]] = row["group"]
    return new_table

def show_results2(results, num_groups):
    print_table = PrettyTable()
    field_names = ["W\P"]
    for player in range(1, num_players + 1):
        field_names.append(str(player))
    print_table.field_names = field_names
    for week in range(1, num_weeks + 1):
        row = [str(week)]
        for player in range(1, num_players + 1):
            row.append(str(results[week][player]))
        print_table.add_row(row)
    print_to_console_and_log(print_table)

def interrupt(s): s.interrupt()

def write_to_cnf(num_vars, num_clauses, problem_name):
    # Create the directory if it doesn't exist
    input_path = online_path + "input_cnf"
    if not os.path.exists(input_path): os.makedirs(input_path)

    # Create the full path to the file "{problem}.cnf" in the directory "input_cnf"
    file_name = problem_name + ".cnf"
    file_path = os.path.join(input_path, file_name)

    # Write data to the file
    with open(file_path, 'w') as writer:
        # Write a line of information about the number of variables and constraints
        writer.write("p cnf " + str(num_vars) + " " + str(num_clauses) + "\n")

        # Write each clause to the file
        for clause in all_clauses:
            for literal in clause: writer.write(str(literal) + " ")
            writer.write("0\n")

    print_to_console_and_log("CNF written to " + file_path + ".\n")

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = online_path + 'out'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print_to_console_and_log(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")

def check_legit(solution, num_groups):
    results = []
    for v in solution:
        if abs(v) > num_players * num_groups * num_weeks: break
        if v > 0 and v <= num_players * num_groups * num_weeks:
            player, group, week = resolve_variable(v, num_groups)
            results.append({"player": player, "group": group, "week": week})

    final_result = process_results(results, num_groups)
    show_results(final_result, num_groups)

    board = process_results2(results, num_groups)
    show_results2(board, num_groups)

    # print_to_console_and_log("Checking validation of the solution...")
    # if (not validate_result(solution)):
    #     print_to_console_and_log("Invalid solution. TERMINATE right now.\n")
    #     return False
    # else: print_to_console_and_log("Valid solution.\n")
    return True

def run_kissat(problem_name):
    # Create the directory if it doesn't exist
    input_path = online_path + "output_kissat"
    if not os.path.exists(input_path): os.makedirs(input_path)

    # Create the full path to the file "{problem}.txt"
    file_name = problem_name + ".txt"
    file_path = os.path.join(input_path, file_name)

    print_to_console_and_log("Running KiSSAT...")
    bashCommand = f"ls input_cnf/{problem_name}.cnf | xargs -n 1 ./kissat --time={time_budget} --relaxed > output_kissat/{problem_name}.txt"
    os.system(bashCommand)
    print_to_console_and_log("KiSSAT finished.")

# solve the problem using the SAT Solver and write the results to xlsx file
def solve_sat_problem():
    global num_players, id_variable, sat_solver, id_counter


    valid_combinations = find_all_valid_combinations()
    for k1, k2, m1, m2, num_groups in valid_combinations:
        id_variable = num_players * num_groups * num_weeks
        id_counter += 1
        # print(f"Trying k1 = {k1}, k2 = {k2}, m1 = {m1}, m2 = {m2}, num_groups = {num_groups}...")
        result_dict = {
            "ID": id_counter,
            "Problem": f"{num_players}-{players_per_group}-{num_weeks}-[{m1}-{m2}]",
            "Type": "nsc",
            "Time": "",
            "Result": "",
            "Variables": 0,
            "Clauses": 0
        }

        print_to_console_and_log(
            f"Problem no. {id_counter}:\n" +
            f"Number of players: {num_players}.\n" +
            f"Number of groups: {num_groups}.\n" +
            f"Players per group: {players_per_group}.\n" +
            f"k1: {k1}.\n" +
            f"k2: {k2}.\n" +
            f"m1: {m1}.\n" +
            f"m2: {m2}.\n" +
            f"Number of weeks: {num_weeks}.\n")

        assert num_groups > 1 and players_per_group[0] > 1

        sat_solver = Glucose3(use_timer = True)
        generate_all_clauses(m1, m2, num_groups)
        # print(sat_solver.get_model(), sat_solver.nof_clauses(), sat_solver.nof_vars())

        # Store the number of variables and clauses before solving the problem
        problem_name = f"{num_players}-{num_groups}-{players_per_group}-{num_weeks}"
        if not enable_kissat:
            num_vars = sat_solver.nof_vars()
            num_clauses = sat_solver.nof_clauses()
        else:
            num_vars = id_variable
            assert num_vars == sat_solver.nof_vars()
            num_clauses = len(all_clauses)
            # print_to_console_and_log(f"{num_clauses} {sat_solver.nof_clauses()}")

        result_dict["Variables"] = num_vars
        result_dict["Clauses"] = num_clauses
        if show_additional_info:
            print_to_console_and_log("Variables: " + str(num_vars))
            print_to_console_and_log("Clauses: " + str(num_clauses))

        print_to_console_and_log("Searching for a solution...")
        timer = Timer(time_budget, interrupt, [sat_solver])
        timer.start()

        sat_status = sat_solver.solve_limited(expect_interrupt = True)
        # print(sat_solver.get_model(), sat_solver.nof_clauses(), sat_solver.nof_vars())
        if sat_status is False:
            elapsed_time = format(sat_solver.time(), ".3f")
            print_to_console_and_log(f"UNSAT. Time run: {elapsed_time}s.\n")
            result_dict["Result"] = "unsat"
            result_dict["Time"] = elapsed_time
            # print(sat_solver.get_model(), sat_solver.nof_clauses(), sat_solver.nof_vars())

        else:
            solution = sat_solver.get_model()
            # print(sat_solver.get_model(), sat_solver.nof_clauses(), sat_solver.nof_vars())
            if solution is None:
                print_to_console_and_log(f"Time limit exceeded ({time_budget}s).\n")
                result_dict["Result"] = "timeout"
                result_dict["Time"] = time_budget

            else:
                elapsed_time = format(sat_solver.time(), ".3f")
                print_to_console_and_log(f"A solution was found in {elapsed_time}s.")
                result_dict["Result"] = "sat"
                result_dict["Time"] = elapsed_time

                if show_additional_info:
                    sat_accum_stats = sat_solver.accum_stats()
                    print_to_console_and_log("Restarts: " + str(sat_accum_stats['restarts']) +
                            ", decisions: " + str(sat_accum_stats['decisions']) +
                            ", propagations: " + str(sat_accum_stats["propagations"]) + '\n')
                if not check_legit(solution, num_groups):
                    timer.cancel()
                    sys.exit(1)

        timer.cancel()
        sat_solver.delete()

        if enable_kissat:
            write_to_cnf(num_vars, num_clauses, problem_name)
            run_kissat(problem_name)
        write_to_xlsx(result_dict)
        all_clauses.clear()

        print_to_console_and_log('-' * 120)

# Open the log file in append mode
log_file = open(online_path + 'console.log', 'a')

# Define a custom print function that writes to both console and log file
def print_to_console_and_log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file = log_file, **kwargs)
    log_file.flush()

# read input data from file data.txt (many lines, each line is number of weeks, number of players per group, number of groups)
# solve the problem
def run_from_input_file():
    global num_players, players_per_group, num_weeks
    with open('data_new.txt') as f:
        for line in f:
            parts = line.split()
            try:
                num_players = int(parts[0])

                # Clean and validate the list part
                list_part = parts[1].strip()
                if not (list_part.startswith('[') and list_part.endswith(']')):
                    raise SyntaxError("List part is not properly formatted")

                players_per_group = ast.literal_eval(list_part)
                num_weeks = int(parts[2])
                solve_sat_problem()
            except (ValueError, SyntaxError) as e:
                print(f"Error parsing line: {line.strip()}")
                print(f"Exception: {e}")
                continue

    log_file.close()


if __name__ == "__main__": run_from_input_file()