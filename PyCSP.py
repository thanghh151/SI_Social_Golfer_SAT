from pycsp3 import *
from itertools import combinations
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
from openpyxl import Workbook
import os
from datetime import datetime
import concurrent.futures
import multiprocessing
from multiprocessing import Value


def solve_social_golfers(data, result_dict, clauses):
    clear()
    nGroups, size, nWeeks = data
    nPlayers = nGroups * size

    print(f"Social Golfer Problem with {nPlayers} players, {nGroups} groups of size {size} and {nWeeks} weeks")

    # x[w][p] is the group admitting on week w the player p
    x = VarArray(size=[nWeeks, nPlayers], dom=range(nGroups))

    satisfy(
        # ensuring that two players don't meet more than one time
        [
            If(
                x[w1][p1] == x[w1][p2],
                Then=x[w2][p1] != x[w2][p2]
            ) for w1, w2 in combinations(range(nWeeks), 2) for p1, p2 in combinations(range(nPlayers), 2)
        ],
        # respecting the size of the groups
        [Cardinality(x[w], occurrences={i: size for i in range(nGroups)}) for w in range(nWeeks)],
        # tag(symmetry-breaking)
        LexIncreasing(x, matrix=True)
    )
    
    # solve_time = time.time() - start_time
    clauses.value = result_dict["Clauses"] = len(posted())
    start_time = time.time()
    solve_result = solve()
    solve_time = time.time() - start_time
    if solve_result is SAT:
        results = []
        for w in range(nWeeks):
            print("Groups of week ", w , [[p for p in range(nPlayers) if x[w][p].value == g] for g in range(nGroups)])
            for g in range(nGroups):
                results.append({"week": w, "group": g, "players": [p for p in range(nPlayers) if x[w][p].value == g]})
        print(f"Constraints: {nGroups * (nWeeks-1) * size}; Variables: {nWeeks * nPlayers}; Time: {solve_time:.4f} seconds")
        result_dict["Result"] = "sat"
    else:
        result_dict["Result"] = "unsat"
    result_dict["Time"] = '{0:.3f}'.format(solve_time)
    write_results_to_excel(result_dict)

def write_results_to_excel(results):
    excel_results = []
    excel_results.append(results)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"out/results_{current_date}.xlsx"

    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()

        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')

        sheet = book['Results']

        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        book.save(excel_file_path)

    else:
        df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print("Result written to Excel file:", os.path.abspath(excel_file_path))
    print("Result added to Excel file.")

id_counter = 1

if __name__ == "__main__":
    # Read data from file
    data_list = []
    with open("data.txt") as file:
        for line in file:
            if line[0] != "#":
                data_list.append([int(x) for x in line.split()])

    # Solve for each set of data
    for data in data_list:
        nGroups, size, nWeeks, = data
        nPlayers = nGroups * size
        clauses = Value('i', 0)
        result_dict = {
            "ID": id_counter,
            "Problem": f"{nGroups}-{size}-{nWeeks}",
            "Type": "PyCSP",
            "Time": "",
            "Result": "",
            "Variables": 0,
            "Clauses": 0
        }
        id_counter += 1
        result_dict["Variables"] = nWeeks * nPlayers
        result_dict["Clauses"] = clauses.value
        # Create a Process
        p = multiprocessing.Process(target=solve_social_golfers, args=(data, result_dict, clauses,))
        p.start()

        # Wait for 600 seconds or until process finishes
        p.join(600)

        # If thread is still active
        if p.is_alive():
            result_dict["Result"] = "timeout"
            result_dict["Time"] = "timeout"
            result_dict["Clauses"] = clauses.value
            write_results_to_excel(result_dict)
            print("solve() function took too long to complete... let's kill it...")
            p.terminate()
            print("Process killed.")
        else:
            print("solve() completed")